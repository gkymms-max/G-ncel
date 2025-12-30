from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Literal
import uuid
from datetime import datetime, timezone, timedelta
import jwt
from passlib.context import CryptContext
import base64
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from urllib.request import urlopen

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
SECRET_KEY = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
security = HTTPBearer()

app = FastAPI()
api_router = APIRouter(prefix="/api")

# Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    password_hash: str
    role: Literal["admin", "user"] = "user"
    permissions: dict = Field(default_factory=lambda: {
        "customers": {"view": False, "create": False, "update": False, "delete": False},
        "suppliers": {"view": False, "create": False, "update": False, "delete": False},
        "products": {"view": False, "create": False, "update": False, "delete": False},
        "quotes": {"view": False, "create": False, "update": False, "delete": False, "approve": False},
        "invoices": {"view": False, "create": False, "update": False, "delete": False, "approve": False},
        "payments": {"view": False, "create": False, "update": False, "delete": False},
        "accounts": {"view": False, "create": False, "update": False, "delete": False},
        "reports": {"view": False, "export": False},
        "settings": {"update": False},
        "categories": {"manage": False},
        "users": {"manage": False},
        "market": {"view": False}
    })
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserLogin(BaseModel):
    username: str
    password: str

class UserRegister(BaseModel):
    username: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    role: str

class Category(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CategoryCreate(BaseModel):
    name: str

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    username: str
    role: str
    created_at: datetime

class Role(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    display_name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RoleCreate(BaseModel):
    name: str
    display_name: str

class ContactChannel(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type: Literal["whatsapp", "instagram", "facebook", "tiktok", "website"]
    title: str  # Görünen isim
    value: str  # Numara, URL, email
    icon: Optional[str] = None  # Icon class
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ContactChannelCreate(BaseModel):
    type: Literal["whatsapp", "instagram", "facebook", "tiktok", "website"]
    title: str
    value: str
    icon: Optional[str] = None

# ============= MUHASEBE MODELLERİ =============

# Tedarikçi Modeli
class Supplier(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    company: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    tax_number: Optional[str] = None
    tax_office: Optional[str] = None
    user_id: str
    balance: float = 0.0  # Borç bakiyesi
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SupplierCreate(BaseModel):
    name: str
    company: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    tax_number: Optional[str] = None
    tax_office: Optional[str] = None

# Fatura Modeli (Alış/Satış)
class InvoiceItem(BaseModel):
    product_id: str
    product_name: str
    unit: str
    quantity: float
    unit_price: float
    vat_rate: float
    vat_amount: float
    subtotal: float
    total: float

class Invoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_number: str
    invoice_type: Literal["sales", "purchase"]  # Satış/Alış
    user_id: str
    # İlgili Taraf (Müşteri veya Tedarikçi)
    party_id: str  # customer_id veya supplier_id
    party_name: str
    party_type: Literal["customer", "supplier"]
    party_tax_number: Optional[str] = None
    party_address: Optional[str] = None
    # Fatura Detayları
    items: List[InvoiceItem]
    subtotal: float
    discount_amount: float = 0
    vat_amount: float
    withholding_amount: float = 0  # Stopaj tutarı
    total: float
    currency: str = "TRY"
    # Tarihler
    invoice_date: datetime
    due_date: datetime
    # Durum
    payment_status: Literal["unpaid", "partial", "paid"] = "unpaid"
    paid_amount: float = 0
    remaining_amount: float = 0
    status: Literal["draft", "pending", "approved", "rejected"] = "draft"
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InvoiceCreate(BaseModel):
    invoice_type: Literal["sales", "purchase"]
    party_id: str
    party_name: str
    party_type: Literal["customer", "supplier"]
    party_tax_number: Optional[str] = None
    party_address: Optional[str] = None
    items: List[InvoiceItem]
    discount_amount: float = 0
    withholding_amount: float = 0
    currency: str = "TRY"
    invoice_date: str
    due_date: str
    notes: Optional[str] = None
    status: Literal["draft", "pending", "approved", "rejected"] = "draft"

# Ödeme/Tahsilat Modeli
class Payment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    payment_type: Literal["income", "expense"]  # Tahsilat/Ödeme
    invoice_id: Optional[str] = None  # Faturaya bağlı ise
    party_id: str  # customer_id veya supplier_id
    party_name: str
    party_type: Literal["customer", "supplier"]
    amount: float
    payment_method: Literal["cash", "bank_transfer", "credit_card", "check", "promissory_note"]
    payment_date: datetime
    account_id: Optional[str] = None  # Kasa/Banka hesabı
    account_name: Optional[str] = None
    description: Optional[str] = None
    check_number: Optional[str] = None  # Çek numarası
    check_date: Optional[datetime] = None  # Çek tarihi
    user_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PaymentCreate(BaseModel):
    payment_type: Literal["income", "expense"]
    invoice_id: Optional[str] = None
    party_id: str
    party_name: str
    party_type: Literal["customer", "supplier"]
    amount: float
    payment_method: Literal["cash", "bank_transfer", "credit_card", "check", "promissory_note"]
    payment_date: str
    account_id: Optional[str] = None
    account_name: Optional[str] = None
    description: Optional[str] = None
    check_number: Optional[str] = None
    check_date: Optional[str] = None

# Kasa/Banka Hesabı Modeli
class Account(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    account_type: Literal["cash", "bank"]
    name: str
    bank_name: Optional[str] = None
    account_number: Optional[str] = None
    iban: Optional[str] = None
    currency: str = "TRY"
    balance: float = 0.0
    user_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AccountCreate(BaseModel):
    account_type: Literal["cash", "bank"]
    name: str
    bank_name: Optional[str] = None
    account_number: Optional[str] = None
    iban: Optional[str] = None
    currency: str = "TRY"
    balance: float = 0.0

# Çek/Senet Modeli
class CheckSenet(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    document_type: Literal["check", "promissory_note"]  # Çek/Senet
    direction: Literal["incoming", "outgoing"]  # Alınan/Verilen
    number: str
    amount: float
    currency: str = "TRY"
    issue_date: datetime
    due_date: datetime
    bank_name: Optional[str] = None
    party_id: str  # customer_id veya supplier_id
    party_name: str
    party_type: Literal["customer", "supplier"]
    status: Literal["pending", "cashed", "bounced", "cancelled"] = "pending"
    payment_id: Optional[str] = None  # İlgili ödeme kaydı
    notes: Optional[str] = None
    user_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CheckSenetCreate(BaseModel):
    document_type: Literal["check", "promissory_note"]
    direction: Literal["incoming", "outgoing"]
    number: str
    amount: float
    currency: str = "TRY"
    issue_date: str
    due_date: str
    bank_name: Optional[str] = None
    party_id: str
    party_name: str
    party_type: Literal["customer", "supplier"]
    notes: Optional[str] = None

# Stok Hareketi Modeli
class StockMovement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_id: str
    product_name: str
    movement_type: Literal["in", "out"]  # Giriş/Çıkış
    quantity: float
    unit: str
    unit_price: float
    total_value: float
    reference_type: Optional[Literal["invoice", "manual", "adjustment"]] = "manual"
    reference_id: Optional[str] = None  # invoice_id vs.
    movement_date: datetime
    warehouse: Optional[str] = None
    notes: Optional[str] = None
    user_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StockMovementCreate(BaseModel):
    product_id: str
    product_name: str
    movement_type: Literal["in", "out"]
    quantity: float
    unit: str
    unit_price: float
    reference_type: Optional[Literal["invoice", "manual", "adjustment"]] = "manual"
    reference_id: Optional[str] = None
    movement_date: str
    warehouse: Optional[str] = None
    notes: Optional[str] = None

# ============= MUHASEBE MODELLERİ SONU =============

class Customer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    company: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    tax_number: Optional[str] = None
    user_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CustomerCreate(BaseModel):
    name: str
    company: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    tax_number: Optional[str] = None


class Product(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    code: str
    name: str
    category: str
    unit: Literal["KG", "Metre", "m²", "Adet"]
    unit_price: float
    currency: str = "EUR"
    package_kg: Optional[float] = None
    package_m2: Optional[float] = None
    package_length: Optional[float] = None
    package_count: Optional[int] = None
    description: Optional[str] = None
    image: Optional[str] = None  # base64 encoded
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProductCreate(BaseModel):
    code: str
    name: str
    category: str
    unit: Literal["KG", "Metre", "m²", "Adet"]
    unit_price: float
    currency: str = "EUR"
    package_kg: Optional[float] = None
    package_m2: Optional[float] = None
    package_length: Optional[float] = None
    package_count: Optional[int] = None
    description: Optional[str] = None
    image: Optional[str] = None

class QuoteItem(BaseModel):
    product_id: str
    product_name: str
    product_code: str
    product_image: Optional[str] = None
    unit: str
    quantity: float
    unit_price: float
    subtotal: float
    note: Optional[str] = None

class Quote(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    quote_number: str
    quote_date: datetime
    validity_date: datetime
    customer_name: str
    customer_email: str
    customer_phone: Optional[str] = None
    customer_id: Optional[str] = None
    customer_company: Optional[str] = None
    currency: str
    items: List[QuoteItem]
    subtotal: float
    discount_type: Literal["percentage", "amount"]
    discount_value: float
    discount_amount: float
    vat_rate: float
    vat_amount: float
    vat_type: Optional[Literal["included", "excluded"]] = "excluded"
    total: float
    notes: Optional[str] = None
    status: Literal["draft", "pending", "approved", "rejected"] = "draft"
    approved_by: Optional[str] = None
    approved_at: Optional[datetime] = None
    rejection_reason: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class QuoteCreate(BaseModel):
    quote_date: str
    validity_date: str
    customer_name: str
    customer_email: str
    customer_phone: Optional[str] = None
    customer_id: Optional[str] = None
    customer_company: Optional[str] = None
    currency: str
    items: List[QuoteItem]
    discount_type: Literal["percentage", "amount"]
    discount_value: float
    vat_rate: float
    vat_type: Optional[Literal["included", "excluded"]] = "excluded"
    notes: Optional[str] = None
    status: Literal["draft", "pending", "approved", "rejected"] = "draft"

class Settings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    company_name: str
    company_address: Optional[str] = None
    company_phone: Optional[str] = None
    company_email: Optional[str] = None
    company_website: Optional[str] = None
    logo: Optional[str] = None
    default_currency: str = "EUR"
    default_vat_rate: float = 18.0
    # PDF Settings
    pdf_show_product_code: bool = False
    pdf_show_unit: bool = False
    pdf_theme: Literal["blue", "green", "purple", "orange"] = "blue"
    # UI Theme
    ui_theme: Literal["light", "dark"] = "light"
    theme_color: Optional[str] = "#4F46E5"
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SettingsUpdate(BaseModel):
    company_name: Optional[str] = None
    company_address: Optional[str] = None
    company_phone: Optional[str] = None
    company_email: Optional[str] = None
    company_website: Optional[str] = None
    logo: Optional[str] = None
    default_currency: Optional[str] = None
    default_vat_rate: Optional[float] = None
    # PDF Settings
    pdf_show_product_code: Optional[bool] = None
    pdf_show_unit: Optional[bool] = None
    pdf_theme: Optional[Literal["blue", "green", "purple", "orange"]] = None
    # UI Theme
    ui_theme: Optional[Literal["light", "dark"]] = None
    theme_color: Optional[str] = None

# Auth helpers
def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(days=7)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user = await db.users.find_one({"username": username}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        
        return {"id": user.get("id"), "username": username, "role": user.get("role", "user")}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_admin_user(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

# Auth endpoints
@api_router.post("/auth/register", response_model=Token)
async def register(user: UserRegister):
    existing_user = await db.users.find_one({"username": user.username}, {"_id": 0})
    if existing_user:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Check if this is the first user
    user_count = await db.users.count_documents({})
    role = "admin" if user_count == 0 else "user"
    
    user_obj = User(
        username=user.username,
        password_hash=hash_password(user.password),
        role=role
    )
    doc = user_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    
    access_token = create_access_token(data={"sub": user.username})
    return Token(access_token=access_token, token_type="bearer", role=role)

@api_router.post("/auth/login", response_model=Token)
async def login(user: UserLogin):
    db_user = await db.users.find_one({"username": user.username}, {"_id": 0})
    if not db_user or not verify_password(user.password, db_user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    
    access_token = create_access_token(data={"sub": user.username})
    return Token(access_token=access_token, token_type="bearer", role=db_user.get("role", "user"))


# Role endpoints (Admin only)
@api_router.get("/roles", response_model=List[Role])
async def get_roles(current_user: dict = Depends(get_admin_user)):
    roles = await db.roles.find({}, {"_id": 0}).to_list(1000)
    for role in roles:
        if isinstance(role['created_at'], str):
            role['created_at'] = datetime.fromisoformat(role['created_at'])
    return roles

@api_router.post("/roles", response_model=Role)
async def create_role(role: RoleCreate, current_user: dict = Depends(get_admin_user)):
    # Check if role name already exists
    existing = await db.roles.find_one({"name": role.name}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Role name already exists")
    
    role_obj = Role(**role.model_dump())
    doc = role_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.roles.insert_one(doc)
    return role_obj

@api_router.delete("/roles/{role_id}")
async def delete_role(role_id: str, current_user: dict = Depends(get_admin_user)):
    # Don't allow deletion of default roles
    role = await db.roles.find_one({"id": role_id}, {"_id": 0})
    if role and role.get('name') in ['admin', 'user']:
        raise HTTPException(status_code=400, detail="Cannot delete default roles")
    
    result = await db.roles.delete_one({"id": role_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Role not found")
    return {"message": "Role deleted successfully"}

# Category endpoints (Admin only)
@api_router.get("/categories", response_model=List[Category])
async def get_categories(current_user: dict = Depends(get_current_user)):
    categories = await db.categories.find({}, {"_id": 0}).to_list(1000)
    for cat in categories:
        if isinstance(cat['created_at'], str):
            cat['created_at'] = datetime.fromisoformat(cat['created_at'])
    return categories

@api_router.post("/categories", response_model=Category)
async def create_category(category: CategoryCreate, current_user: dict = Depends(get_admin_user)):
    # Check if category already exists
    existing = await db.categories.find_one({"name": category.name}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Category already exists")
    
    cat_obj = Category(name=category.name)
    doc = cat_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.categories.insert_one(doc)
    return cat_obj

@api_router.put("/categories/{category_id}", response_model=Category)
async def update_category(category_id: str, category: CategoryCreate, current_user: dict = Depends(get_admin_user)):
    existing = await db.categories.find_one({"id": category_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Category not found")
    
    # Check if new name already exists (excluding current category)
    name_exists = await db.categories.find_one({"name": category.name, "id": {"$ne": category_id}}, {"_id": 0})
    if name_exists:
        raise HTTPException(status_code=400, detail="Category name already exists")
    
    await db.categories.update_one({"id": category_id}, {"$set": {"name": category.name}})
    updated = await db.categories.find_one({"id": category_id}, {"_id": 0})
    if isinstance(updated['created_at'], str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return updated

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, current_user: dict = Depends(get_admin_user)):
    result = await db.categories.delete_one({"id": category_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    return {"message": "Category deleted successfully"}

class UserCreateByAdmin(BaseModel):
    username: str
    password: str
    role: Literal["admin", "user"] = "user"

# Contact Channels endpoints (Admin only)
@api_router.get("/contact-channels", response_model=List[ContactChannel])
async def get_contact_channels(current_user: dict = Depends(get_current_user)):
    channels = await db.contact_channels.find({}, {"_id": 0}).to_list(1000)
    for channel in channels:
        if isinstance(channel['created_at'], str):
            channel['created_at'] = datetime.fromisoformat(channel['created_at'])
    return channels

@api_router.post("/contact-channels", response_model=ContactChannel)
async def create_contact_channel(channel: ContactChannelCreate, current_user: dict = Depends(get_admin_user)):
    channel_obj = ContactChannel(**channel.model_dump())
    doc = channel_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.contact_channels.insert_one(doc)
    return channel_obj

@api_router.delete("/contact-channels/{channel_id}")
async def delete_contact_channel(channel_id: str, current_user: dict = Depends(get_admin_user)):
    result = await db.contact_channels.delete_one({"id": channel_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contact channel not found")
    return {"message": "Contact channel deleted successfully"}
@api_router.get("/users", response_model=List[UserResponse])
async def get_users(current_user: dict = Depends(get_admin_user)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    for user in users:
        if isinstance(user['created_at'], str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
    return users

@api_router.post("/users", response_model=UserResponse)
async def create_user_by_admin(user: UserCreateByAdmin, current_user: dict = Depends(get_admin_user)):
    existing = await db.users.find_one({"username": user.username}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    user_obj = User(
        username=user.username,
        password_hash=hash_password(user.password),
        role=user.role
    )
    doc = user_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    
    return UserResponse(
        id=user_obj.id,
        username=user_obj.username,
        role=user_obj.role,
        created_at=user_obj.created_at
    )

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_admin_user)):
    # Don't allow deleting yourself
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user and user["username"] == current_user["username"]:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted successfully"}

@api_router.patch("/users/{user_id}/permissions")
async def update_user_permissions(
    user_id: str,
    permissions: dict,
    current_user: dict = Depends(get_admin_user)
):
    """Update user permissions - Admin only"""
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"permissions": permissions}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if isinstance(user.get('created_at'), str):
        user['created_at'] = datetime.fromisoformat(user['created_at'])
    return user

@api_router.put("/users/{user_id}/password")
async def reset_user_password(user_id: str, password_data: dict, current_user: dict = Depends(get_admin_user)):
    new_password = password_data.get("password")
    if not new_password or len(new_password) < 4:
        raise HTTPException(status_code=400, detail="Password must be at least 4 characters")
    
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"password_hash": hash_password(new_password)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "Password reset successfully"}

@api_router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, role_data: dict, current_user: dict = Depends(get_admin_user)):
    new_role = role_data.get("role")
    if new_role not in ["admin", "user"]:
        raise HTTPException(status_code=400, detail="Invalid role. Must be 'admin' or 'user'")
    
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"role": new_role}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "Role updated successfully"}


# Product endpoints
@api_router.post("/products", response_model=Product)
async def create_product(product: ProductCreate, current_user: dict = Depends(get_current_user)):
    product_obj = Product(**product.model_dump())
    doc = product_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.products.insert_one(doc)
    return product_obj


# Customer endpoints
@api_router.get("/customers", response_model=List[Customer])
async def get_customers(current_user: dict = Depends(get_current_user)):
    customers = await db.customers.find({"user_id": current_user['id']}, {"_id": 0}).to_list(1000)
    for customer in customers:
        if isinstance(customer['created_at'], str):
            customer['created_at'] = datetime.fromisoformat(customer['created_at'])
    return customers

@api_router.post("/customers", response_model=Customer)
async def create_customer(customer: CustomerCreate, current_user: dict = Depends(get_current_user)):
    customer_obj = Customer(**customer.model_dump(), user_id=current_user['id'])
    doc = customer_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.customers.insert_one(doc)
    return customer_obj

@api_router.put("/customers/{customer_id}", response_model=Customer)
async def update_customer(customer_id: str, customer: CustomerCreate, current_user: dict = Depends(get_current_user)):
    result = await db.customers.update_one(
        {"id": customer_id, "user_id": current_user['id']},
        {"$set": customer.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    updated = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if isinstance(updated['created_at'], str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Customer(**updated)

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.customers.delete_one({"id": customer_id, "user_id": current_user['id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"message": "Customer deleted successfully"}

# ============= TEDARİKÇİ ENDPOINTS =============

@api_router.get("/suppliers")
async def get_suppliers(current_user: dict = Depends(get_current_user)):
    suppliers = await db.suppliers.find({"user_id": current_user["username"]}, {"_id": 0}).to_list(1000)
    for supplier in suppliers:
        if isinstance(supplier.get('created_at'), str):
            supplier['created_at'] = datetime.fromisoformat(supplier['created_at'])
    return suppliers

@api_router.post("/suppliers")
async def create_supplier(supplier_data: SupplierCreate, current_user: dict = Depends(get_current_user)):
    supplier = Supplier(**supplier_data.model_dump(), user_id=current_user["username"])
    supplier_dict = supplier.model_dump()
    supplier_dict['created_at'] = supplier_dict['created_at'].isoformat()
    await db.suppliers.insert_one(supplier_dict)
    return supplier

@api_router.put("/suppliers/{supplier_id}")
async def update_supplier(supplier_id: str, supplier_data: SupplierCreate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in supplier_data.model_dump().items() if v is not None}
    result = await db.suppliers.update_one(
        {"id": supplier_id, "user_id": current_user["username"]},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if isinstance(supplier['created_at'], str):
        supplier['created_at'] = datetime.fromisoformat(supplier['created_at'])
    return supplier

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.suppliers.delete_one({"id": supplier_id, "user_id": current_user["username"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return {"message": "Supplier deleted successfully"}

# ============= HESAP (KASA/BANKA) ENDPOINTS =============

@api_router.get("/accounts")
async def get_accounts(current_user: dict = Depends(get_current_user)):
    accounts = await db.accounts.find({"user_id": current_user["username"]}, {"_id": 0}).to_list(1000)
    for account in accounts:
        if isinstance(account.get('created_at'), str):
            account['created_at'] = datetime.fromisoformat(account['created_at'])
    return accounts

@api_router.post("/accounts")
async def create_account(account_data: AccountCreate, current_user: dict = Depends(get_current_user)):
    account = Account(**account_data.model_dump(), user_id=current_user["username"])
    account_dict = account.model_dump()
    account_dict['created_at'] = account_dict['created_at'].isoformat()
    await db.accounts.insert_one(account_dict)
    return account

@api_router.put("/accounts/{account_id}")
async def update_account(account_id: str, account_data: AccountCreate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in account_data.model_dump().items() if v is not None and k != 'balance'}
    result = await db.accounts.update_one(
        {"id": account_id, "user_id": current_user["username"]},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Account not found")
    
    account = await db.accounts.find_one({"id": account_id}, {"_id": 0})
    if isinstance(account['created_at'], str):
        account['created_at'] = datetime.fromisoformat(account['created_at'])
    return account

@api_router.delete("/accounts/{account_id}")
async def delete_account(account_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.accounts.delete_one({"id": account_id, "user_id": current_user["username"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Account not found")
    return {"message": "Account deleted successfully"}

# Existing Products endpoints continue below...

# ============= FATURA ENDPOINTS =============

@api_router.get("/invoices")
async def get_invoices(
    invoice_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"user_id": current_user["username"]}
    if invoice_type:
        query["invoice_type"] = invoice_type
    
    invoices = await db.invoices.find(query, {"_id": 0}).to_list(1000)
    for invoice in invoices:
        for field in ['invoice_date', 'due_date', 'created_at']:
            if field in invoice and isinstance(invoice[field], str):
                invoice[field] = datetime.fromisoformat(invoice[field])
    return invoices

@api_router.post("/invoices")
async def create_invoice(invoice_data: InvoiceCreate, current_user: dict = Depends(get_current_user)):
    # Calculate totals
    subtotal = sum(item.subtotal for item in invoice_data.items)
    vat_amount = sum(item.vat_amount for item in invoice_data.items)
    total = subtotal + vat_amount - invoice_data.discount_amount - invoice_data.withholding_amount
    
    # Generate invoice number
    invoice_count = await db.invoices.count_documents({"user_id": current_user["username"]})
    invoice_number = f"FTR-{datetime.now().year}-{invoice_count + 1:05d}"
    
    invoice = Invoice(
        invoice_number=invoice_number,
        user_id=current_user["username"],
        invoice_date=datetime.fromisoformat(invoice_data.invoice_date),
        due_date=datetime.fromisoformat(invoice_data.due_date),
        subtotal=subtotal,
        vat_amount=vat_amount,
        total=total,
        remaining_amount=total,
        **{k: v for k, v in invoice_data.model_dump().items() if k not in ['invoice_date', 'due_date']}
    )
    
    invoice_dict = invoice.model_dump()
    invoice_dict['invoice_date'] = invoice_dict['invoice_date'].isoformat()
    invoice_dict['due_date'] = invoice_dict['due_date'].isoformat()
    invoice_dict['created_at'] = invoice_dict['created_at'].isoformat()
    
    await db.invoices.insert_one(invoice_dict)
    
    # Update party balance
    collection = db.customers if invoice.party_type == "customer" else db.suppliers
    await collection.update_one(
        {"id": invoice.party_id},
        {"$inc": {"balance": total if invoice.invoice_type == "sales" else -total}}
    )
    
    # Create stock movements
    for item in invoice.items:
        movement = StockMovement(
            product_id=item.product_id,
            product_name=item.product_name,
            movement_type="out" if invoice.invoice_type == "sales" else "in",
            quantity=item.quantity,
            unit=item.unit,
            unit_price=item.unit_price,
            total_value=item.total,
            reference_type="invoice",
            reference_id=invoice.id,
            movement_date=invoice.invoice_date,
            user_id=current_user["username"]
        )
        movement_dict = movement.model_dump()
        movement_dict['movement_date'] = movement_dict['movement_date'].isoformat()
        movement_dict['created_at'] = movement_dict['created_at'].isoformat()
        await db.stock_movements.insert_one(movement_dict)
    
    return invoice

@api_router.patch("/invoices/{invoice_id}/status")
async def update_invoice_status(
    invoice_id: str,
    status: Literal["draft", "pending", "approved", "rejected"],
    current_user: dict = Depends(get_current_user)
):
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {"status": status}}
    )
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    for field in ['invoice_date', 'due_date', 'created_at']:
        if field in invoice and isinstance(invoice[field], str):
            invoice[field] = datetime.fromisoformat(invoice[field])
    return invoice

@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Revert balance changes
    collection = db.customers if invoice['party_type'] == "customer" else db.suppliers
    multiplier = 1 if invoice['invoice_type'] == "sales" else -1
    await collection.update_one(
        {"id": invoice['party_id']},
        {"$inc": {"balance": -invoice['total'] * multiplier}}
    )
    
    # Delete related stock movements
    await db.stock_movements.delete_many({"reference_id": invoice_id})
    
    await db.invoices.delete_one({"id": invoice_id})
    return {"message": "Invoice deleted successfully"}

# ============= ÖDEME/TAHSİLAT ENDPOINTS =============

@api_router.get("/payments")
async def get_payments(
    payment_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"user_id": current_user["username"]}
    if payment_type:
        query["payment_type"] = payment_type
    
    payments = await db.payments.find(query, {"_id": 0}).to_list(1000)
    for payment in payments:
        for field in ['payment_date', 'check_date', 'created_at']:
            if field in payment and payment[field] and isinstance(payment[field], str):
                payment[field] = datetime.fromisoformat(payment[field])
    return payments

@api_router.post("/payments")
async def create_payment(payment_data: PaymentCreate, current_user: dict = Depends(get_current_user)):
    payment = Payment(
        user_id=current_user["username"],
        payment_date=datetime.fromisoformat(payment_data.payment_date),
        check_date=datetime.fromisoformat(payment_data.check_date) if payment_data.check_date else None,
        **{k: v for k, v in payment_data.model_dump().items() if k not in ['payment_date', 'check_date']}
    )
    
    payment_dict = payment.model_dump()
    payment_dict['payment_date'] = payment_dict['payment_date'].isoformat()
    if payment_dict.get('check_date'):
        payment_dict['check_date'] = payment_dict['check_date'].isoformat()
    payment_dict['created_at'] = payment_dict['created_at'].isoformat()
    
    await db.payments.insert_one(payment_dict)
    
    # Update invoice if linked
    if payment.invoice_id:
        invoice = await db.invoices.find_one({"id": payment.invoice_id}, {"_id": 0})
        if invoice:
            new_paid = invoice.get('paid_amount', 0) + payment.amount
            new_remaining = invoice['total'] - new_paid
            new_status = "paid" if new_remaining <= 0 else "partial"
            
            await db.invoices.update_one(
                {"id": payment.invoice_id},
                {"$set": {
                    "paid_amount": new_paid,
                    "remaining_amount": new_remaining,
                    "payment_status": new_status
                }}
            )
    
    # Update account balance
    if payment.account_id:
        multiplier = 1 if payment.payment_type == "income" else -1
        await db.accounts.update_one(
            {"id": payment.account_id},
            {"$inc": {"balance": payment.amount * multiplier}}
        )
    
    # Update party balance
    collection = db.customers if payment.party_type == "customer" else db.suppliers
    multiplier = -1 if payment.payment_type == "income" else 1
    await collection.update_one(
        {"id": payment.party_id},
        {"$inc": {"balance": payment.amount * multiplier}}
    )
    
    return payment

@api_router.delete("/payments/{payment_id}")
async def delete_payment(payment_id: str, current_user: dict = Depends(get_current_user)):
    payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    # Revert invoice payment
    if payment.get('invoice_id'):
        invoice = await db.invoices.find_one({"id": payment['invoice_id']}, {"_id": 0})
        if invoice:
            new_paid = max(0, invoice.get('paid_amount', 0) - payment['amount'])
            new_remaining = invoice['total'] - new_paid
            new_status = "unpaid" if new_paid <= 0 else ("paid" if new_remaining <= 0 else "partial")
            
            await db.invoices.update_one(
                {"id": payment['invoice_id']},
                {"$set": {
                    "paid_amount": new_paid,
                    "remaining_amount": new_remaining,
                    "payment_status": new_status
                }}
            )
    
    # Revert account balance
    if payment.get('account_id'):
        multiplier = 1 if payment['payment_type'] == "income" else -1
        await db.accounts.update_one(
            {"id": payment['account_id']},
            {"$inc": {"balance": -payment['amount'] * multiplier}}
        )
    
    # Revert party balance
    collection = db.customers if payment['party_type'] == "customer" else db.suppliers
    multiplier = -1 if payment['payment_type'] == "income" else 1
    await collection.update_one(
        {"id": payment['party_id']},
        {"$inc": {"balance": -payment['amount'] * multiplier}}
    )
    
    await db.payments.delete_one({"id": payment_id})
    return {"message": "Payment deleted successfully"}

# ============= ÇEK/SENET & STOK ENDPOINTS =============

@api_router.get("/checks")
async def get_checks(current_user: dict = Depends(get_current_user)):
    checks = await db.checks.find({"user_id": current_user["username"]}, {"_id": 0}).to_list(1000)
    for check in checks:
        for field in ['issue_date', 'due_date', 'created_at']:
            if field in check and isinstance(check[field], str):
                check[field] = datetime.fromisoformat(check[field])
    return checks

@api_router.post("/checks")
async def create_check(check_data: CheckSenetCreate, current_user: dict = Depends(get_current_user)):
    check = CheckSenet(
        user_id=current_user["username"],
        issue_date=datetime.fromisoformat(check_data.issue_date),
        due_date=datetime.fromisoformat(check_data.due_date),
        **{k: v for k, v in check_data.model_dump().items() if k not in ['issue_date', 'due_date']}
    )
    
    check_dict = check.model_dump()
    check_dict['issue_date'] = check_dict['issue_date'].isoformat()
    check_dict['due_date'] = check_dict['due_date'].isoformat()
    check_dict['created_at'] = check_dict['created_at'].isoformat()
    
    await db.checks.insert_one(check_dict)
    return check

@api_router.get("/stock-movements")
async def get_stock_movements(current_user: dict = Depends(get_current_user)):
    movements = await db.stock_movements.find({"user_id": current_user["username"]}, {"_id": 0}).to_list(1000)
    for movement in movements:
        for field in ['movement_date', 'created_at']:
            if field in movement and isinstance(movement[field], str):
                movement[field] = datetime.fromisoformat(movement[field])
    return movements

@api_router.post("/stock-movements")
async def create_stock_movement(movement_data: StockMovementCreate, current_user: dict = Depends(get_current_user)):
    total_value = movement_data.quantity * movement_data.unit_price
    
    movement = StockMovement(
        user_id=current_user["username"],
        movement_date=datetime.fromisoformat(movement_data.movement_date),
        total_value=total_value,
        **{k: v for k, v in movement_data.model_dump().items() if k != 'movement_date'}
    )
    
    movement_dict = movement.model_dump()
    movement_dict['movement_date'] = movement_dict['movement_date'].isoformat()
    movement_dict['created_at'] = movement_dict['created_at'].isoformat()
    
    await db.stock_movements.insert_one(movement_dict)
    return movement

# Existing Products endpoints continue below...

@api_router.get("/products", response_model=List[Product])
async def get_products(current_user: dict = Depends(get_current_user)):
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    for product in products:
        if isinstance(product['created_at'], str):
            product['created_at'] = datetime.fromisoformat(product['created_at'])
    return products

@api_router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str, current_user: dict = Depends(get_current_user)):
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    if isinstance(product['created_at'], str):
        product['created_at'] = datetime.fromisoformat(product['created_at'])
    return product

@api_router.put("/products/{product_id}", response_model=Product)
async def update_product(product_id: str, product: ProductCreate, current_user: dict = Depends(get_current_user)):
    existing_product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not existing_product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    update_data = product.model_dump()
    await db.products.update_one({"id": product_id}, {"$set": update_data})
    
    updated_product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if isinstance(updated_product['created_at'], str):
        updated_product['created_at'] = datetime.fromisoformat(updated_product['created_at'])
    return updated_product

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product deleted successfully"}

# Quote endpoints
@api_router.post("/quotes", response_model=Quote)
async def create_quote(quote: QuoteCreate, current_user: dict = Depends(get_current_user)):
    # Get last quote number for this user
    last_quote = await db.quotes.find_one({"user_id": current_user["username"]}, {"_id": 0, "quote_number": 1}, sort=[("created_at", -1)])
    if last_quote and last_quote.get("quote_number"):
        last_num = int(last_quote["quote_number"].split("-")[-1])
        quote_number = f"FT-{last_num + 1:05d}"
    else:
        quote_number = "FT-00001"
    
    # Calculate totals
    subtotal = sum(item.subtotal for item in quote.items)
    
    if quote.discount_type == "percentage":
        discount_amount = subtotal * (quote.discount_value / 100)
    else:
        discount_amount = quote.discount_value
    
    subtotal_after_discount = subtotal - discount_amount
    vat_amount = subtotal_after_discount * (quote.vat_rate / 100)
    total = subtotal_after_discount + vat_amount
    
    quote_obj = Quote(
        user_id=current_user["username"],
        quote_number=quote_number,
        quote_date=datetime.fromisoformat(quote.quote_date),
        validity_date=datetime.fromisoformat(quote.validity_date),
        customer_name=quote.customer_name,
        customer_email=quote.customer_email,
        customer_phone=quote.customer_phone,
        currency=quote.currency,
        items=quote.items,
        subtotal=subtotal,
        discount_type=quote.discount_type,
        discount_value=quote.discount_value,
        discount_amount=discount_amount,
        vat_rate=quote.vat_rate,
        vat_amount=vat_amount,
        total=total,
        notes=quote.notes
    )
    
    doc = quote_obj.model_dump()
    doc['quote_date'] = doc['quote_date'].isoformat()
    doc['validity_date'] = doc['validity_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.quotes.insert_one(doc)
    return quote_obj

@api_router.get("/quotes", response_model=List[Quote])
async def get_quotes(current_user: dict = Depends(get_current_user)):
    quotes = await db.quotes.find({"user_id": current_user["username"]}, {"_id": 0}).to_list(1000)
    for quote in quotes:
        if isinstance(quote['quote_date'], str):
            quote['quote_date'] = datetime.fromisoformat(quote['quote_date'])
        if isinstance(quote['validity_date'], str):
            quote['validity_date'] = datetime.fromisoformat(quote['validity_date'])
        if isinstance(quote['created_at'], str):
            quote['created_at'] = datetime.fromisoformat(quote['created_at'])
    return quotes

@api_router.get("/quotes/{quote_id}", response_model=Quote)
async def get_quote(quote_id: str, current_user: dict = Depends(get_current_user)):
    quote = await db.quotes.find_one({"id": quote_id, "user_id": current_user["username"]}, {"_id": 0})
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    if isinstance(quote['quote_date'], str):
        quote['quote_date'] = datetime.fromisoformat(quote['quote_date'])
    if isinstance(quote['validity_date'], str):
        quote['validity_date'] = datetime.fromisoformat(quote['validity_date'])
    if isinstance(quote['created_at'], str):
        quote['created_at'] = datetime.fromisoformat(quote['created_at'])
    return quote

@api_router.delete("/quotes/{quote_id}")
async def delete_quote(quote_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.quotes.delete_one({"id": quote_id, "user_id": current_user["username"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Quote not found")
    return {"message": "Quote deleted successfully"}

@api_router.patch("/quotes/{quote_id}/status")
async def update_quote_status(
    quote_id: str, 
    status: Literal["draft", "pending", "approved", "rejected"],
    rejection_reason: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    quote = await db.quotes.find_one({"id": quote_id}, {"_id": 0})
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    update_data = {"status": status}
    
    if status == "approved":
        update_data["approved_by"] = current_user["username"]
        update_data["approved_at"] = datetime.now(timezone.utc).isoformat()
    elif status == "rejected":
        if rejection_reason:
            update_data["rejection_reason"] = rejection_reason
    
    await db.quotes.update_one(
        {"id": quote_id},
        {"$set": update_data}
    )
    
    updated_quote = await db.quotes.find_one({"id": quote_id}, {"_id": 0})
    if isinstance(updated_quote['quote_date'], str):
        updated_quote['quote_date'] = datetime.fromisoformat(updated_quote['quote_date'])
    if isinstance(updated_quote['validity_date'], str):
        updated_quote['validity_date'] = datetime.fromisoformat(updated_quote['validity_date'])
    if isinstance(updated_quote['created_at'], str):
        updated_quote['created_at'] = datetime.fromisoformat(updated_quote['created_at'])
    
    return updated_quote

@api_router.get("/quotes/pending/list")
async def get_pending_quotes(current_user: dict = Depends(get_current_user)):
    """Get all pending quotes (admin only)"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    quotes = await db.quotes.find({"status": "pending"}, {"_id": 0}).to_list(1000)
    for quote in quotes:
        if isinstance(quote['quote_date'], str):
            quote['quote_date'] = datetime.fromisoformat(quote['quote_date'])
        if isinstance(quote['validity_date'], str):
            quote['validity_date'] = datetime.fromisoformat(quote['validity_date'])
        if isinstance(quote['created_at'], str):
            quote['created_at'] = datetime.fromisoformat(quote['created_at'])
    
    return quotes

@api_router.get("/quotes/{quote_id}/pdf")
async def get_quote_pdf(quote_id: str, current_user: dict = Depends(get_current_user)):
    quote = await db.quotes.find_one({"id": quote_id, "user_id": current_user["username"]}, {"_id": 0})
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    settings = await db.settings.find_one({"user_id": current_user["username"]}, {"_id": 0})
    
    # Convert datetime strings
    if isinstance(quote['quote_date'], str):
        quote['quote_date'] = datetime.fromisoformat(quote['quote_date'])
    if isinstance(quote['validity_date'], str):
        quote['validity_date'] = datetime.fromisoformat(quote['validity_date'])
    
    # Generate PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
    story = []
    
    # Register Turkish-compatible fonts
    try:
        pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
        font_name = 'DejaVuSans'
        font_bold = 'DejaVuSans-Bold'
    except:
        # Fallback to Helvetica if DejaVu not available
        font_name = 'Helvetica'
        font_bold = 'Helvetica-Bold'
    
    # Get theme color from settings
    theme_color = settings.get('theme_color', '#4F46E5') if settings else '#4F46E5'
    theme = {"primary": theme_color, "secondary": theme_color}
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontName=font_bold, fontSize=20, textColor=colors.HexColor(theme["primary"]), spaceAfter=12)
    heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontName=font_bold, fontSize=12, textColor=colors.HexColor('#374151'), spaceAfter=6)
    normal_style = ParagraphStyle('CustomNormal', parent=styles['Normal'], fontName=font_name, fontSize=9, textColor=colors.HexColor('#4b5563'))
    
    # Logo and company info header
    header_data = []
    
    # Company info section
    if settings:
        company_info = []
        if settings.get('company_name'):
            company_info.append([Paragraph(settings['company_name'], title_style)])
        if settings.get('company_address'):
            company_info.append([Paragraph(settings['company_address'], normal_style)])
        
        contact_info = []
        if settings.get('company_phone'):
            contact_info.append(f"Tel: {settings['company_phone']}")
        if settings.get('company_email'):
            contact_info.append(f"E-posta: {settings['company_email']}")
        if settings.get('company_website'):
            contact_info.append(f"Web: {settings['company_website']}")
        
        if contact_info:
            company_info.append([Paragraph(" | ".join(contact_info), normal_style)])
        
        # Create header layout
        if settings.get('logo'):
            try:
                logo_data = base64.b64decode(settings['logo'].split(',')[1] if ',' in settings['logo'] else settings['logo'])
                logo_img = RLImage(io.BytesIO(logo_data), width=4*cm, height=3*cm)
                
                # Create header table with logo and company info side by side
                header_table_data = [[logo_img, company_info]]
                header_table = Table(header_table_data, colWidths=[5*cm, 11*cm])
                header_table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                    ('ALIGN', (1, 0), (1, 0), 'LEFT'),
                    ('LEFTPADDING', (1, 0), (1, 0), 20),
                ]))
                story.append(header_table)
            except:
                # If logo fails, just add company info
                for info in company_info:
                    story.append(info[0])
        else:
            # No logo, just company info
            for info in company_info:
                story.append(info[0])
    else:
        story.append(Paragraph("Firma Adı", title_style))
    
    story.append(Spacer(1, 0.8*cm))
    
    # Quote title
    quote_title = Paragraph("FİYAT TEKLİFİ", 
                           ParagraphStyle('QuoteTitle', parent=title_style, 
                                        fontSize=24, textColor=colors.HexColor(theme["primary"]), 
                                        alignment=1, spaceAfter=20))
    story.append(quote_title)
    story.append(Spacer(1, 0.5*cm))
    
    # Quote info
    info_data = [
        ["Teklif No:", quote['quote_number'], "Müşteri:", quote['customer_name']],
        ["Tarih:", quote['quote_date'].strftime('%d.%m.%Y'), "E-posta:", quote['customer_email']],
        ["Geçerlilik:", quote['validity_date'].strftime('%d.%m.%Y'), "Telefon:", quote.get('customer_phone', '-')]
    ]
    
    info_table = Table(info_data, colWidths=[3*cm, 5*cm, 3*cm, 5*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#6b7280')),
        ('TEXTCOLOR', (2, 0), (2, -1), colors.HexColor('#6b7280')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#111827')),
        ('TEXTCOLOR', (3, 0), (3, -1), colors.HexColor('#111827')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.5*cm))
    
    # Items table - first get all products for lookup
    all_products = await db.products.find({}, {"_id": 0}).to_list(1000)
    product_lookup = {p['id']: p for p in all_products}
    
    table_data = [['Ürün Adı', 'Birim', 'Koli/PK', 'Birim Fiyat', 'Miktar', 'Tutar']]
    for item in quote['items']:
        # Get product details
        product = product_lookup.get(item['product_id'])
        
        # Calculate package info and actual quantity
        package_info = "-"
        actual_quantity = item['quantity']
        unit = item['unit']
        
        if product:
            # Determine package size based on unit
            if unit == "KG" and product.get('package_kg'):
                package_info = f"{product['package_kg']}"
            elif unit == "m²" and product.get('package_m2'):
                package_info = f"{product['package_m2']}"
            elif unit == "Metre" and product.get('package_length'):
                package_info = f"{product['package_length']}"
            elif unit == "Adet" and product.get('package_count'):
                package_info = f"{product['package_count']}"
        
        # Format quantity display
        if item.get('display_text'):
            # If we have display text, extract the calculated amount
            # Example: "2 paket (100 m²)" -> show "100"
            if '(' in item['display_text'] and ')' in item['display_text']:
                calc_part = item['display_text'].split('(')[1].split(')')[0]
                actual_quantity = calc_part.split(' ')[0]
            else:
                actual_quantity = str(item['quantity'])
        else:
            actual_quantity = str(item['quantity'])
        
        table_data.append([
            item['product_name'],  # Full product name, no truncation
            unit,
            package_info,
            f"{item['unit_price']:.2f}",
            actual_quantity,
            f"{item['subtotal']:.2f} {quote['currency']}"
        ])
    
    items_table = Table(table_data, colWidths=[7*cm, 1.8*cm, 1.8*cm, 2.2*cm, 2.2*cm, 2.5*cm])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(theme["primary"])),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),   # Birim column centered
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),   # Koli/PK column centered
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),    # Birim Fiyat column right
        ('ALIGN', (4, 0), (4, -1), 'CENTER'),   # Miktar column centered
        ('ALIGN', (5, 0), (5, -1), 'RIGHT'),    # Tutar column right aligned
        ('FONTNAME', (0, 0), (-1, 0), font_bold),
        ('FONTSIZE', (0, 0), (-1, 0), 8),       # Header font size
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 1), (-1, -1), 7),      # Content font size
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        # Allow text wrapping in product name column for long names
        ('WORDWRAP', (0, 1), (0, -1), 'CJK'),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 0.3*cm))
    
    # Totals
    totals_data = [
        ['Ara Toplam:', f"{quote['subtotal']:.2f} {quote['currency']}"]
    ]
    
    # İndirim varsa ekle
    if quote['discount_amount'] > 0:
        totals_data.append(['İndirim:', f"-{quote['discount_amount']:.2f} {quote['currency']}"])
    
    # KDV varsa ekle
    if quote['vat_amount'] > 0:
        totals_data.append(['KDV (%{:.0f})'.format(quote['vat_rate']), f"{quote['vat_amount']:.2f} {quote['currency']}"])
    
    totals_data.append(['GENEL TOPLAM:', f"{quote['total']:.2f} {quote['currency']}"])
    
    totals_table = Table(totals_data, colWidths=[13*cm, 3*cm])
    totals_table.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 2), font_name),
        ('FONTSIZE', (0, 0), (-1, 2), 9),
        ('FONTNAME', (0, 3), (-1, 3), font_bold),
        ('FONTSIZE', (0, 3), (-1, 3), 11),
        ('TEXTCOLOR', (0, 3), (-1, 3), colors.HexColor(theme["primary"])),
        ('LINEABOVE', (0, 3), (-1, 3), 1.5, colors.HexColor(theme["primary"])),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(totals_table)
    
    if quote.get('notes'):
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph("<b>Notlar:</b>", heading_style))
        story.append(Paragraph(quote['notes'], normal_style))
    
    doc.build(story)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=teklif_{quote['quote_number']}.pdf"}
    )

# Settings endpoints
@api_router.get("/settings", response_model=Settings)
async def get_settings(current_user: dict = Depends(get_current_user)):
    settings = await db.settings.find_one({"user_id": current_user["username"]}, {"_id": 0})
    if not settings:
        # Create default settings for this user
        default_settings = Settings(
            id=str(uuid.uuid4()),
            user_id=current_user["username"],
            company_name="Firma Adı",
            company_address="Firma Adresi",
            company_phone="+90 XXX XXX XX XX",
            company_email="info@firma.com",
            company_website="www.firma.com"
        )
        doc = default_settings.model_dump()
        doc['updated_at'] = doc['updated_at'].isoformat()
        await db.settings.insert_one(doc)
        return default_settings
    
    if isinstance(settings['updated_at'], str):
        settings['updated_at'] = datetime.fromisoformat(settings['updated_at'])
    return settings

@api_router.put("/settings", response_model=Settings)
async def update_settings(settings_update: SettingsUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in settings_update.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    # Check if settings exist for this user
    existing = await db.settings.find_one({"user_id": current_user["username"]}, {"_id": 0})
    if not existing:
        # Create new settings for this user
        update_data['id'] = str(uuid.uuid4())
        update_data['user_id'] = current_user["username"]
        if 'company_name' not in update_data:
            update_data['company_name'] = "Firma Adı"
        await db.settings.insert_one(update_data)
    else:
        await db.settings.update_one(
            {"user_id": current_user["username"]},
            {"$set": update_data}
        )
    
    settings = await db.settings.find_one({"user_id": current_user["username"]}, {"_id": 0})
    if isinstance(settings['updated_at'], str):
        settings['updated_at'] = datetime.fromisoformat(settings['updated_at'])
    return settings

# Excel Export endpoints
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

@api_router.get("/export/products")
async def export_products_excel(current_user: dict = Depends(get_current_user)):
    """Export all products to Excel"""
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ürünler"
    
    # Headers
    headers = ["Kod", "Ürün Adı", "Kategori", "Birim", "Birim Fiyat", "Para Birimi", "Açıklama"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for product in products:
        ws.append([
            product.get('code', ''),
            product.get('name', ''),
            product.get('category', ''),
            product.get('unit', ''),
            product.get('unit_price', 0),
            product.get('currency', 'EUR'),
            product.get('description', '')
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=urunler.xlsx"}
    )

@api_router.get("/export/customers")
async def export_customers_excel(current_user: dict = Depends(get_current_user)):
    """Export all customers to Excel"""
    customers = await db.customers.find({"user_id": current_user["username"]}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Müşteriler"
    
    # Headers
    headers = ["Ad Soyad", "Firma", "E-posta", "Telefon", "Adres", "Vergi No"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for customer in customers:
        ws.append([
            customer.get('name', ''),
            customer.get('company', ''),
            customer.get('email', ''),
            customer.get('phone', ''),
            customer.get('address', ''),
            customer.get('tax_number', '')
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=musteriler.xlsx"}
    )

@api_router.get("/export/quotes")
async def export_quotes_excel(current_user: dict = Depends(get_current_user)):
    """Export all quotes to Excel"""
    quotes = await db.quotes.find({"user_id": current_user["username"]}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Teklifler"
    
    # Headers
    headers = ["Teklif No", "Müşteri", "Tarih", "Geçerlilik", "Durum", "Para Birimi", "Ara Toplam", "KDV", "Toplam"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for quote in quotes:
        quote_date = quote.get('quote_date')
        if isinstance(quote_date, str):
            quote_date = datetime.fromisoformat(quote_date)
        
        validity_date = quote.get('validity_date')
        if isinstance(validity_date, str):
            validity_date = datetime.fromisoformat(validity_date)
        
        status_map = {
            "draft": "Taslak",
            "pending": "Beklemede",
            "approved": "Onaylandı",
            "rejected": "Reddedildi"
        }
        
        ws.append([
            quote.get('quote_number', ''),
            quote.get('customer_name', ''),
            quote_date.strftime('%d.%m.%Y') if quote_date else '',
            validity_date.strftime('%d.%m.%Y') if validity_date else '',
            status_map.get(quote.get('status', 'draft'), 'Taslak'),
            quote.get('currency', 'EUR'),
            round(quote.get('subtotal', 0), 2),
            round(quote.get('vat_amount', 0), 2),
            round(quote.get('total', 0), 2)
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=teklifler.xlsx"}
    )

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()